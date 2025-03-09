"""Excel parse plugin"""

import os
import itertools
from io import FileIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as SheetImage
from server.file import FileItem, get_md5_from_bytes
from .config import ReadXLSXConfig
from ..core import DataParser
from ..types import PaginationConfig, CanPaginationData, BasicValueType, FileValue
from ..exceptions import InvalidConfigValue
from ..utils import data_cache


def parse_cell(cell: Cell | ReadOnlyCell) -> BasicValueType:
    """Parse the cell value"""
    if not isinstance(cell, ReadOnlyCell):
        link = cell.hyperlink
        if not link is None:
            return {
                "url": str(link.target),
                "text": str(link.display),
                "type": "url",
            }
    value = cell.value
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return int(round(value.timestamp() * 1000))


def validate_read_config(config: ReadXLSXConfig):
    """Validate the config"""
    sheet_name = config.get("sheet_name")
    if sheet_name is None:
        raise InvalidConfigValue("`sheet_name` is required")
    sheet_name = str(sheet_name)
    data_range = config.get("data_range", None)
    if not data_range is None and not isinstance(data_range, str):
        data_range = str(data_range)
    header_index = config.get("header", None)
    if header_index is None:
        header_index = 1
    performance_mode = config.get("performance_mode", False)
    if performance_mode is None:
        performance_mode = False
    return (sheet_name, data_range, header_index, performance_mode)


def close_wb_file(wb: Workbook, f: FileIO):
    """Get the close function for workbook and file"""

    def close():
        """Close the workbook and file"""
        wb.close()
        f.close()

    return close


def get_workbook(
    file: FileItem,
    read_only: bool = False,
):
    """Get the workbook object from the file

    Args:
        file (FileItem): file

    Returns:
        (Workbook, Callable[[], None]): workbook object and close function
    """
    f = open(file.file_path, "rb")
    wb = load_workbook(
        filename=f,
        read_only=read_only,
        keep_vba=True,
        data_only=True,
        rich_text=False,
        keep_links=True,
    )
    return (wb, close_wb_file(wb, f))


def validate_data(data: list[list[BasicValueType]]):
    """Check if the file can be parsed"""
    header_checker = [not cell is None for cell in data[0]]
    if not all(header_checker):
        return {
            "error": "The header row contains empty cells or merged cells.",
            "can_parse": False,
        }
    return {
        "can_parse": True,
    }


def get_xlsx_iter_data(
    data: FileItem,
    config: ReadXLSXConfig,
):
    sheet_name = config.get("sheet_name")
    if sheet_name is None:
        raise InvalidConfigValue("`sheet_name` is required")
    sheet_name = str(sheet_name)
    data_range = config.get("data_range", None)
    if not data_range is None and not isinstance(data_range, str):
        data_range = str(data_range)
    header_index = config.get("header", None)
    if header_index is None:
        header_index = 1
    if not isinstance(header_index, int):
        header_index = int(header_index)
    wb, close = get_workbook(data)
    ws = wb[sheet_name]
    _min_row = ws.min_row
    _max_row = ws.max_row

    if data_range:
        ws: tuple[tuple[Cell]] = ws[data_range]
        min_row = max(header_index, 0)
        if header_index >= len(ws):
            raise InvalidConfigValue(f"Header index {header_index} is out of range.")
        _data = ([cell.value for cell in row] for row in ws[min_row:])
    else:
        min_row = max(header_index, _min_row)
        if min_row > _max_row:
            raise InvalidConfigValue(f"Page token is out of range.")
        max_row = min(min_row, _max_row)
        if header_index > max_row:
            raise InvalidConfigValue(f"Header index {header_index} is out of range.")
        _data = (
            [cell.value for cell in row]
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
            )
        )
    return (_data, close)


def get_xlsx_info(data: FileItem):
    """Get the information of the excel file"""
    wb, close = get_workbook(data)
    sheet_names = wb.sheetnames
    default_sheet_name = sheet_names[0]
    ws = wb[default_sheet_name]
    header = paginate_load_xlsx(
        data,
        {
            "page_size": 1,
            "page_token": 0,
            "config": {
                "sheet_name": default_sheet_name,
                "header": 1,
                "performance_mode": True,
            }
        }
    )
    close()


def parse_data(data: FileItem, config: ReadXLSXConfig, context: DataParser):
    sheet_name = config.get("sheet_name")
    if sheet_name is None:
        raise InvalidConfigValue("`sheet_name` is required")
    sheet_name = str(sheet_name)
    data_range = config.get("data_range", None)
    if not data_range is None and not isinstance(data_range, str):
        data_range = str(data_range)
    header_index = config.get("header", None)
    if header_index is None:
        header_index = 1
    if not isinstance(header_index, int):
        header_index = int(header_index)
    wb, close = get_workbook(data)


def load_image(image: SheetImage, save_dir: str) -> tuple[str, FileValue]:
    anchor = (
        f"{get_column_letter(image.anchor._from.col + 1)}{image.anchor._from.row + 1}"
    )
    _image = Image.open(image.ref).convert("RGB")
    md5 = get_md5_from_bytes(_image.tobytes())
    name = f"{md5}.{image.format}"
    file_path = os.path.join(save_dir, name)
    if not os.path.exists(file_path):
        _image.save(file_path)
    return (
        anchor,
        {
            "name": name,
            "type": "file",
            "size": os.path.getsize(file_path),
            "md5": md5,
            "token": None,
        },
    )


def load_images(ws: Worksheet, f: FileItem):
    _images: list[SheetImage] = ws._images
    image_path = os.path.join(f.dir_path, "attachments")
    images: dict[str, list[FileValue]] = {}
    if not os.path.exists(image_path) and len(_images) > 0:
        os.makedirs(image_path)
    for i in _images:
        anchor, file_value = load_image(i, image_path)
        if not anchor in images:
            images[anchor] = []
        images[anchor].append(file_value)
    return images


def thread_load_images(ws: Worksheet, f: FileItem, max_workers: int = 20):
    _images = ws._images
    image_path = os.path.join(f.dir_path, "attachments")
    images: dict[str, list[FileValue]] = {}
    if not os.path.exists(image_path) and len(_images) > 0:
        os.makedirs(image_path)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        res = executor.map(
            load_image, [i for i in _images], itertools.repeat(image_path)
        )
    for anchor, file_value in res:
        if not anchor in images:
            images[anchor] = []
        images[anchor].append(file_value)
    return images


def get_preview_cache_key(config: PaginationConfig[ReadXLSXConfig]):
    _config = config["config"]
    return f"SN{_config["sheet_name"]}DR{_config.get("data_range", None)}H{_config.get("header", 1)}S{config["page_size"]}T{config.get("page_token", 0)}"


@data_cache[PaginationConfig[ReadXLSXConfig]](get_cache_key=get_preview_cache_key)
def paginate_load_xlsx(data: FileItem, config: PaginationConfig[ReadXLSXConfig]):
    """Preview excel file"""
    page_size = config.get("page_size", None)
    if page_size is None:
        raise InvalidConfigValue("`page_size` is required")
    page_size = int(page_size)
    page_token = config.get("page_token", None)
    if page_token is None:
        page_token = 0
    if not isinstance(page_token, int):
        page_token = int(page_token)

    _config = config["config"]
    if _config is None:
        raise InvalidConfigValue("`config` is required")
    sheet_name, data_range, header_index, performance_mode = validate_read_config(
        _config
    )

    wb, close = get_workbook(data, read_only=performance_mode)
    ws = wb[sheet_name]
    _min_row = ws.min_row
    _max_row = ws.max_row
    _min_col = ws.min_column
    _max_col = ws.max_column
    has_more = True
    if not performance_mode:
        images = thread_load_images(ws, data)

    if data_range:
        ws = ws[data_range]
        min_row = max(header_index + page_token * page_size - 1, 0)
        if min_row > len(ws):
            raise InvalidConfigValue(f"Page token {page_token} is out of range.")
        max_row = min(len(ws), min_row + page_size)
        if header_index > max_row:
            raise InvalidConfigValue(f"Header index {header_index} is out of range.")
        has_more = max_row < len(ws)
        ws = ws[min_row : max_row + 1]
        if performance_mode or len(images) == 0:
            _data = [[parse_cell(cell) for cell in row] for row in ws]
        else:
            _data = [
                [
                    (
                        images[f"{get_column_letter(cell.column)}{cell.row}"]
                        if f"{get_column_letter(cell.column)}{cell.row}" in images
                        else parse_cell(cell)
                    )
                    for cell in row
                ]
                for row in ws
            ]
    else:
        min_row = max(header_index + page_token * page_size, _min_row)
        if min_row > _max_row:
            raise InvalidConfigValue(f"Page token {page_token} is out of range.")
        max_row = min(min_row + page_size - 1, _max_row)
        if header_index > max_row:
            raise InvalidConfigValue(f"Header index {header_index} is out of range.")
        if performance_mode or len(images) == 0:
            _data = [
                [parse_cell(cell) for cell in row]
                for row in ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=_min_col,
                    max_col=_max_col,
                )
            ]
        else:
            _data = [
                [
                    (
                        images[f"{get_column_letter(cell.column)}{cell.row}"]
                        if f"{get_column_letter(cell.column)}{cell.row}" in images
                        else parse_cell(cell)
                    )
                    for cell in row
                ]
                for row in ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=_min_col,
                    max_col=_max_col,
                )
            ]
    close()
    res: CanPaginationData[list[list[BasicValueType]]] = {
        "data": _data,
        "page_size": page_size,
        "page_token": page_token,
        "has_more": has_more,
    }
    return res
