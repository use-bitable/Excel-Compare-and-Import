import os
from openpyxl import load_workbook
from datetime import time, datetime

TEST_DIR = os.path.dirname(__file__)
XLSX_FILE_PATH = os.path.join(TEST_DIR, "./assets/data.xlsx")

wb = load_workbook(XLSX_FILE_PATH, data_only=True, keep_links=True, keep_vba=True)
DEFAULT_SHEET_NAME = "Sheet1"

ws = wb[DEFAULT_SHEET_NAME]

for row in ws.iter_rows(min_row=2, max_row=5):
    for cell in row:
        print(type(cell.value), cell.value, cell.is_date)
