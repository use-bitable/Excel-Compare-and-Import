"""Excel parse plugin"""

import os
import itertools
from io import FileIO
from typing import DefaultDict
from collections import defaultdict
from datetime import datetime, time, timedelta, date
from concurrent.futures import ThreadPoolExecutor
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as SheetImage
from app.file import FileItem, get_md5_from_bytes
from .constants import (
    DEFAULT_SHEET_NAME,
    DEFAULT_HEADER,
    DEFAULT_PAGE_TOKEN,
    DEFAULT_DATA_RANGE,
    DEFAULT_PERFORMANCE_MODE,
)
from .types import ReadXLSXConfig, DataRange
from ..types import (
    PaginationConfig,
    CanPaginationData,
    BasicValueType,
    FileValue,
    ParsedData,
)
from ..exceptions import InvalidConfigValue, InvalidHeader
from ..utils import data_cache, parse_data_to_dict


def parse_cell(cell: Cell | ReadOnlyCell) -> BasicValueType:
    """Parse the cell value"""
    if not isinstance(cell, ReadOnlyCell):
        link = cell.hyperlink
        if link is not None:
            return {
                "url": str(link.target),
                "text": str(link.display),
                "type": "url",
            }
    value = cell.value
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, time):
        d = date.today()
        dt = datetime.combine(d, value)
        return dt
    if isinstance(value, datetime):
        return value


def parse_data_range(
    data_range: str | DataRange | None,
) -> tuple[int, int, int, int] | tuple[None, None, None, None]:
    """Parse the data range"""
    return (
        DEFAULT_DATA_RANGE
        if data_range is None
        else (
            range_boundaries(data_range)
            if isinstance(data_range, str)
            else (
                data_range.get("min_col"),
                data_range.get("min_row"),
                data_range.get("max_col"),
                data_range.get("max_row"),
            )
        )
    )


def validate_read_config(config: ReadXLSXConfig):
    """Validate the config

    Args:
        config (ReadXLSXConfig): config

    Returns:
        tuple[str | None, tuple[int, int, int, int] | tuple[None, None, None, None], int, bool]: sheet_name, data_range, header_index, performance_mode
    """
    if config is None:
        return (
            DEFAULT_SHEET_NAME,
            DEFAULT_DATA_RANGE,
            DEFAULT_HEADER,
            DEFAULT_PERFORMANCE_MODE,
        )
    sheet_name = config.get("sheet_name")
    if not sheet_name is None:
        sheet_name = str(sheet_name)
    data_range = parse_data_range(config.get("data_range"))
    header_index = config.get("header")
    if header_index is None:
        header_index = DEFAULT_HEADER
    if header_index < 1:
        raise InvalidConfigValue("`header` should >= 1")
    performance_mode = config.get("performance_mode")
    if performance_mode is None:
        performance_mode = DEFAULT_PERFORMANCE_MODE
    return (sheet_name, data_range, header_index, performance_mode)


def close_wb_file(wb: Workbook, f: FileIO):
    """Get the close function for workbook and file"""

    def close():
        """Close the workbook and file"""
        wb.close()
        f.close()

    return close


def get_workbook(
    file: FileItem,
    read_only: bool = False,
):
    """Get the workbook object from the file

    Args:
        file (FileItem): file

    Returns:
        (Workbook, Callable[[], None]): workbook object and close function
    """
    f = open(file.file_path, "rb")
    wb = load_workbook(
        filename=f,
        read_only=read_only,
        keep_vba=True,
        data_only=True,
        rich_text=False,
        keep_links=True,
    )
    return (wb, close_wb_file(wb, f))


def validate_header(data: list[BasicValueType]):
    """Check if the file can be parsed"""
    header_checker = [isinstance(cell, (str, int, float)) for cell in data]
    if not all(header_checker):
        raise InvalidHeader(
            f"Invalid header cell: {[get_column_letter(index + 1) for index, v in enumerate(header_checker) if not v]}"
        )
    return True


def load_image(
    image: SheetImage, save_dir: str, parent_token: str
) -> tuple[str, FileValue]:
    anchor = (
        f"{get_column_letter(image.anchor._from.col + 1)}{image.anchor._from.row + 1}"
    )
    _image = Image.open(image.ref).convert("RGB")
    md5 = get_md5_from_bytes(_image.tobytes())
    name = f"{md5}.{image.format}"
    file_path = os.path.join(save_dir, name)
    if not os.path.exists(file_path):
        _image.save(file_path)
    return (
        anchor,
        {
            "name": name,
            "type": "file",
            "size": os.path.getsize(file_path),
            "md5": md5,
            "token": None,
            "parent_token": parent_token,
        },
    )


def get_image_cache_key(ws: Worksheet, f: FileItem):
    return os.path.join(f.dir_path, "cache", "images_map", f"{ws.title}.json")


@data_cache(get_cache_key=get_image_cache_key)
def load_images(ws: Worksheet, f: FileItem):
    _images: list[SheetImage] = ws._images
    image_path = os.path.join(f.dir_path, "attachments")
    images: DefaultDict[str, list[FileValue]] = defaultdict(list)
    if not os.path.exists(image_path) and len(_images) > 0:
        os.makedirs(image_path)
    if len(_images) > 0:
        for i in _images:
            anchor, file_value = load_image(i, image_path, f.token)
            images[anchor].append(file_value)
    return images


@data_cache(get_cache_key=get_image_cache_key)
def thread_load_images(ws: Worksheet, f: FileItem, max_workers: int = 32):
    _images = ws._images
    image_path = os.path.join(f.dir_path, "attachments")
    images: DefaultDict[str, list[FileValue]] = defaultdict(list)
    if not os.path.exists(image_path) and len(_images) > 0:
        os.makedirs(image_path)
    if len(_images) > 0:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            res = executor.map(
                load_image,
                _images,
                itertools.repeat(image_path),
                itertools.repeat(f.token),
            )
        for anchor, file_value in res:
            images[anchor].append(file_value)
    return images


def iter_row_xlsx(
    file: FileItem,
    config: ReadXLSXConfig | None = None,
):
    """Get the row iterator"""
    sheet_name, data_range, header_index, performance_mode = validate_read_config(
        config
    )

    wb, close = get_workbook(file, read_only=performance_mode)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    _min_col, _min_row, _max_col, _max_row = (
        (
            ws.min_column,
            ws.min_row,
            ws.max_column,
            ws.max_row,
        )
        if all([c is None for c in data_range])
        else data_range
    )
    header = [parse_cell(c) for c in ws[_min_row + header_index - 1]]
    validate_header(header)
    header = [str(cell) for cell in header]
    images = thread_load_images(ws, file) if not performance_mode else {}
    for row in ws.iter_rows(
        min_row=header_index + _min_row - 1,
        max_row=_max_row,
        min_col=_min_col,
        max_col=_max_col,
    ):
        yield dict(
            zip(
                header,
                (
                    [parse_cell(cell) for cell in row]
                    if performance_mode or len(images) == 0
                    else [
                        (
                            images[f"{get_column_letter(cell.column)}{cell.row}"]
                            if f"{get_column_letter(cell.column)}{cell.row}" in images
                            else parse_cell(cell)
                        )
                        for cell in row
                    ]
                ),
            )
        )
    close()


def get_paginate_cache_key(
    file: FileItem,
    config: PaginationConfig[ReadXLSXConfig] | None,
    parse_data: bool = False,
):
    if config is None:
        config = {
            "page_token": DEFAULT_PAGE_TOKEN,
            "page_size": None,
            "config": None,
        }
    _config = config.get("config")
    sheet_name, data_range, header_index, _ = validate_read_config(_config)
    min_col, min_row, max_col, max_row = data_range
    dr_str = f"{min_col}_{min_row}_{max_col}_{max_row}"
    page_size = config.get("page_size")
    page_token = (
        config.get("page_token")
        if not config.get("page_token") is None
        else DEFAULT_PAGE_TOKEN
    )
    cache_filename = f"SN{sheet_name}DR{dr_str}H{header_index}S{page_size}T{page_token}{'_parsed' if parse_data else ''}.json"
    return os.path.join(file.dir_path, "cache", "paginate_data", cache_filename)


DEFAULT_PAGINATE_CONFIG = {
    "page_token": DEFAULT_PAGE_TOKEN,
    "page_size": None,
    "config": None,
}


@data_cache(get_cache_key=get_paginate_cache_key)
def paginate_load_xlsx(
    data: FileItem,
    config: PaginationConfig[ReadXLSXConfig] | None,
    parse_data: bool = False,
) -> CanPaginationData[ParsedData]:
    """Preview excel file"""
    if config is None:
        config = DEFAULT_PAGINATE_CONFIG
    page_size = config.get("page_size")
    if not page_size is None:
        page_size = int(page_size)
    page_token = config.get("page_token")
    if page_token is None:
        page_token = DEFAULT_PAGE_TOKEN
    if page_token < 0:
        raise InvalidConfigValue("`page_token` should >= 0")

    _config = config.get("config")
    sheet_name, data_range, header_index, performance_mode = validate_read_config(
        _config
    )

    wb, close = get_workbook(data, read_only=performance_mode)
    sheet_names = wb.sheetnames
    if sheet_name is None:
        sheet_name = sheet_names[0]
    ws = wb[sheet_name]
    _default_range = [
        ws.min_column,
        ws.min_row,
        ws.max_column,
        ws.max_row,
    ]
    _min_col, _min_row, _max_col, _max_row = (
        v if v is not None else _default_range[i] for i, v in enumerate(data_range)
    )
    has_more = True
    images = thread_load_images(ws, data) if not performance_mode else {}
    min_row = (
        header_index
        + _min_row
        - 1
        + (page_token * page_size if not page_size is None else 0)
        + (1 if page_token == 0 else 0)
    )
    max_row = _max_row if page_size is None else min(_max_row, min_row + page_size - 1)
    if min_row > max_row:
        close()
        raise InvalidConfigValue(f"Header index or page token is out of range.")
    has_more = max_row < _max_row
    header = [
        parse_cell(c)
        for i, c in enumerate(ws[_min_row + header_index - 1])
        if i < _max_col and i >= _min_col - 1
    ]
    _data = (
        [
            [parse_cell(cell) for cell in row]
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=_min_col,
                max_col=_max_col,
            )
        ]
        if performance_mode or len(images) == 0
        else [
            [
                (
                    images[f"{get_column_letter(cell.column)}{cell.row}"]
                    if f"{get_column_letter(cell.column)}{cell.row}" in images
                    else parse_cell(cell)
                )
                for cell in row
            ]
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=_min_col,
                max_col=_max_col,
            )
        ]
    )

    close()
    errors = []
    can_parse = True

    try:
        validate_header(header)
        header = [str(cell) for cell in header]
    except Exception as e:
        can_parse = False
        errors.append(str(e))

    parsed_data = (
        parse_data_to_dict(_data, header) if parse_data and can_parse else None
    )
    total = _max_row - _min_row + 1
    if total <= 1:
        can_parse = False
        errors.append(
            "No data, need at least 2 rows of excel file (one for header, one for data)."
        )
    res_data = (
        _data if not parse_data or not can_parse or parsed_data is None else parsed_data
    )
    res: CanPaginationData[ParsedData] = {
        "data": {
            "data": res_data,
            "meta": {
                "fields": header,
                "total": total,
                "errors": errors,
                "can_parse": can_parse,
                "extra": {
                    "sheet_name": sheet_name,
                    "sheet_names": sheet_names,
                    "data_range": {
                        "min_col": _min_col,
                        "min_row": _min_row,
                        "max_col": _max_col,
                        "max_row": _max_row,
                    },
                    "header_index": header_index,
                },
            },
        },
        "page_size": len(res_data),
        "page_token": page_token,
        "has_more": has_more,
    }
    return res
