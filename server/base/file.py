#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Author: qww
# @Version: 1.0
# @License: MIT
# @Description: File utility functions

from typing import Optional
from requests import get
from requests.exceptions import RequestException
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from baseopensdk.api.drive.v1 import UploadAllMediaRequest, \
  UploadAllMediaRequestBody, UploadAllMediaResponse
from baseopensdk import BaseClient
from openpyxl import load_workbook
from .field import FieldMap
from .const import BASE_FILE_SIZE_LIMIT
from .exception import InvalidFileException
from ..types import HeadersMapping





def download_file(
        url: str,
        headers: Optional[HeadersMapping]=None):
    """Download file from URL.

    Args:
        url (str): URL of the file

    Raises:
        requests: Failed to download file from URL

    Returns:
        FileStorage: File
    """
    res = get(url, headers=headers, stream=True)
    if res.status_code == 200:
        filename = secure_filename(url.split('/')[-1])
        file = FileStorage(
          stream=res.content,
          filename=filename,
          headers=res.headers
        )
        return file
    raise RequestException(f"Failed to download file from {url}: status code {res.status_code}")


def allowed_file(
        allowed_extensions: Optional[list[str]] = None,
        size_limit: Optional[int] = None,
):
    """Create file checker.

    Args:
        allowed_extensions (list[str], optional): The allowed 
        extensions list of file. Defaults to None.
        size_limit (int, optional): The size limit of file. Defaults to None.

    Returns:
        Callable[[FileStorage], bool]: File checker
    """
    allowed_extensions = [ext.lower().replace('.', '') for ext in (allowed_extensions or [])]
    def checker(file: FileStorage):
        """File checker.

        Args:
            file (FileStorage): File

        Returns:
            bool: True if the file is allowed, False otherwise
        """
        if len(allowed_extensions) > 0 and \
          file.filename.split('.')[-1].lower() not in allowed_extensions:
            return False
        if size_limit and file.content_length > size_limit:
            return False
        return True
    return checker


base_file_checker = allowed_file(size_limit=BASE_FILE_SIZE_LIMIT)


def upload_file_to_base(
        file: FileStorage,
        base_id: str,
        base_client: BaseClient
):
    """Upload file to Base.

    Args:
        file (FileStorage): File to upload
        base_id (str): Base ID
        base_client (BaseClient): Base client

    Raises:
        InvalidFileException: File is too large
        requests: Failed to upload file to Base

    Returns:
        str: File token
    """
    if not base_file_checker(file):
        raise InvalidFileException(
            f"Too large file {file.filename}({file.size} bytes) (should < {BASE_FILE_SIZE_LIMIT} bytes)"
        )
    filename = file.filename
    size = file.content_length
    body = UploadAllMediaRequestBody.builder()\
        .file(file.stream)\
        .file_name(filename)\
        .size(size)\
        .parent_node(base_id)\
        .parent_type('bitable_file')\
        .build()
    request = UploadAllMediaRequest.builder()\
        .request_body(body).build()
    res: UploadAllMediaResponse = \
        base_client.drive.v1.media.upload_all(request)
    if res.success():
        return res.data.file_token
    raise RequestException(f"Failed to upload file to Base {base_id}")


xlsx_file_checker = allowed_file(allowed_extensions=['xlsx', 'xls'])


def read_xlsx(
        file: FileStorage,
        sheet_name: str,
        fields_map: list[FieldMap],
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        data_range: str = None
):
    """Read xlsx file."""
    if not xlsx_file_checker(file):
        raise InvalidFileException(f"File {file.filename} is not allowed")
    wb = load_workbook(filename=file.stream, read_only=True)
    ws = wb[sheet_name]
    rows = None
    if data_range:
        rows = ws[data_range].iter_rows()
    else:
        rows = ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col)
    return ws
